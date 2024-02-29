import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# Give the location of the file
path = "/content/input.xlsx"

# Read the Excel file into a DataFrame
df = pd.read_excel(path)

# Fill NaN values forward along columns
df = df.fillna(method='ffill', axis=0)

# Delete the first two columns ( zeydin )
df = df.drop(df.columns[:2], axis=1)

# Rename the columns
df = df.rename(columns={
    df.columns[0]: "periode",
    df.columns[1]: "zone",
    df.columns[2]: "country",
    df.columns[3]: "operator",
    df.columns[4]: "nbrsms",
    df.columns[5]: "nbrappels",
    df.columns[6]: "durappels"
})

# Delete the first 5 rows ( zeydin espaces )
df = df.iloc[5:]

# ne7ina zone international
df = df[df['zone'] != 'International']

# zone specifique mobile or not?
df['is_mobile_zone'] = df['zone'].str.contains('Mobile')

# lem el mobile
df['durappelsmob'] = df.loc[df['operator'].str.contains('-Mob') | df['is_mobile_zone'], 'durappels']

# lem el fixe
df['durappelsfix'] = df.loc[~df['operator'].str.contains('-Mob') & ~df['is_mobile_zone'], 'durappels']

# Group by 'zone' and calculate the sum for each group
result = df.groupby('zone').agg({
    'nbrsms': 'sum',
    'durappelsmob': 'sum',
    'durappelsfix': 'sum'
}).reset_index()

# Create a new DataFrame with the desired format
output_df = pd.DataFrame({
    'Zone': result['zone'],
    'Nombre SMS': result['nbrsms'],
    'Durée (min) Mobile': result['durappelsmob'],
    'Durée (min) Fixe': result['durappelsfix']
})

# Sort the DataFrame by the 'Zone' column
output_df = output_df.sort_values(by='Zone')

# Add the total row
output_df.loc['Total'] = output_df[['Nombre SMS', 'Durée (min) Mobile', 'Durée (min) Fixe']].sum()
output_df.at[output_df.index[-1], 'Zone'] = 'Total:'

# Add an empty row
output_df = output_df.append(pd.Series(), ignore_index=True)

# Find the index of the new empty row
empty_row_index = output_df.index[-1]

# Add the month information in the last row
month_str = pd.to_datetime(result['zone'].apply(lambda x: df[df['zone'] == x]['periode'].values[0]), format='%d/%m/%Y').dt.strftime('%B %Y').values[0]
output_df.at[empty_row_index, 'Zone'] = 'Month: ' 
output_df.at[empty_row_index, 'Nombre SMS'] = str(month_str)



# Write the DataFrame to an Excel file with formatting
output_path = '/content/output.xlsx'
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    output_df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the XlsxWriter workbook and worksheet objects from the writer object.
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Define styles
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    blue_fill = PatternFill(start_color='5b9bd5', end_color='5b9bd5', fill_type='solid')
    # Apply formatting to the first row
    for col_idx in range(1, output_df.shape[1] + 1):
        worksheet.cell(row=1, column=col_idx).fill = blue_fill
        worksheet.cell(row=1, column=col_idx).alignment = center_alignment
        worksheet.cell(row=1, column=col_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    # Apply formatting to the last two rows
    for row_idx in range(output_df.shape[0] - 1, output_df.shape[0] + 1 ):
        for col_idx in range(1, output_df.shape[1] + 1):
            worksheet.cell(row=row_idx + 1, column=col_idx).fill = blue_fill
            worksheet.cell(row=row_idx + 1, column=col_idx).alignment = center_alignment
            worksheet.cell(row=row_idx + 1, column=col_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Apply formatting to all other cells
    for row_idx in range(1, output_df.shape[0] ):
        for col_idx in range(1, output_df.shape[1] + 1):
            worksheet.cell(row=row_idx + 1, column=col_idx).alignment = center_alignment
            worksheet.cell(row=row_idx + 1, column=col_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    last_row_idx = output_df.shape[0]
    last_column_idx = output_df.shape[1]
    worksheet.merge_cells(start_row=last_row_idx+1, start_column=last_column_idx - 2, end_row=last_row_idx+1, end_column=last_column_idx)

print(f"Formatted output saved to {output_path}")